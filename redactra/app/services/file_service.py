import io
import json
import re
import zipfile
from dataclasses import dataclass
from html import escape
from pathlib import Path

from fastapi import UploadFile

from app.core.config import settings


SUPPORTED_TYPES = {".pdf", ".docx", ".sql", ".csv", ".json", ".txt", ".png", ".jpg", ".jpeg", ".xlsx"}
IMAGE_TYPES = {".png", ".jpg", ".jpeg"}

SIZE_THRESHOLDS = {
    ".pdf": (5 * 1024, 50 * 1024 * 1024),
    ".docx": (2 * 1024, 50 * 1024 * 1024),
    ".xlsx": (10, 50 * 1024 * 1024),
    ".sql": (100, 100 * 1024 * 1024),
    ".csv": (10, 100 * 1024 * 1024),
    ".json": (10, 100 * 1024 * 1024),
    ".txt": (1, 10 * 1024 * 1024),
    ".png": (5 * 1024, 20 * 1024 * 1024),
    ".jpg": (5 * 1024, 20 * 1024 * 1024),
    ".jpeg": (5 * 1024, 20 * 1024 * 1024),
}

MAGIC_BYTES = {
    ".pdf": b"%PDF",
    ".docx": b"PK\x03\x04",
    ".xlsx": b"PK\x03\x04",
    ".png": b"\x89PNG",
    ".jpg": b"\xff\xd8\xff",
    ".jpeg": b"\xff\xd8\xff",
}


@dataclass
class OcrToken:
    text: str
    start: int
    end: int
    left: int
    top: int
    width: int
    height: int


def ensure_storage_dirs(raw_path: Path, sanitized_path: Path) -> None:
    raw_path.mkdir(parents=True, exist_ok=True)
    sanitized_path.mkdir(parents=True, exist_ok=True)


async def save_upload(upload: UploadFile, destination: Path) -> None:
    data = await upload.read()
    destination.write_bytes(data)


def validate_file_size(extension: str, size: int) -> None:
    if extension not in SIZE_THRESHOLDS:
        return
    min_b, max_b = SIZE_THRESHOLDS[extension]
    if size < min_b or size > max_b:
        raise ValueError(f"File size anomaly for {extension}: {size} bytes")


def validate_magic_bytes(file_bytes: bytes, extension: str) -> None:
    expected = MAGIC_BYTES.get(extension)
    if not expected:
        return
    if file_bytes.startswith(expected):
        return
    # Some PDFs can have leading whitespace/newlines before the header.
    if extension == ".pdf" and file_bytes[:1024].lstrip().startswith(expected):
        return
    raise ValueError("File header does not match extension")


def strip_exif_if_image(file_bytes: bytes, extension: str) -> tuple[bytes, bool, int]:
    if extension not in IMAGE_TYPES:
        return file_bytes, False, 0

    try:
        from PIL import Image
    except Exception:
        return file_bytes, False, 0

    img = Image.open(io.BytesIO(file_bytes))
    exif = img.getexif()
    exif_count = len(exif) if exif else 0

    clean = Image.new(img.mode, img.size)
    clean.putdata(list(img.getdata()))
    out = io.BytesIO()
    format_hint = "PNG" if extension == ".png" else "JPEG"
    clean.save(out, format=format_hint)
    return out.getvalue(), exif_count > 0, exif_count


def _extract_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        data = zf.read("word/document.xml").decode("utf-8", errors="ignore")
    data = re.sub(r"<w:p[^>]*>", "\n", data)
    data = re.sub(r"<[^>]+>", "", data)
    return re.sub(r"\n+", "\n", data)


def _extract_pdf_text(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except Exception as exc:
        raise RuntimeError("pypdf is required for PDF parsing") from exc

    reader = PdfReader(str(path))
    text = []
    for page in reader.pages:
        text.append(page.extract_text() or "")
    return "\n".join(text)


def _xlsx_to_lines(path: Path) -> tuple[list[str], list[tuple[str, str]]]:
    """
    Flatten an .xlsx into a stable list of string lines (one per string cell).
    Returns (lines, coords) where coords items are (sheet_name, cell_coordinate).
    This is used both for extraction and writing sanitized output back.
    """
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required for XLSX parsing") from exc

    wb = load_workbook(filename=str(path), data_only=True)
    lines: list[str] = []
    coords: list[tuple[str, str]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.strip():
                    # Avoid introducing newlines that break mapping
                    safe = " ".join(v.splitlines()).strip()
                    lines.append(safe)
                    coords.append((ws.title, cell.coordinate))
    return lines, coords


def _extract_xlsx_text(path: Path) -> str:
    lines, _ = _xlsx_to_lines(path)
    return "\n".join(lines)


def _configure_tesseract(pytesseract_module) -> None:
    if settings.tesseract_cmd:
        pytesseract_module.pytesseract.tesseract_cmd = settings.tesseract_cmd


def extract_image_text_with_tokens(path: Path) -> tuple[str, list[OcrToken]]:
    try:
        from PIL import Image
        import pytesseract
        from pytesseract import Output
    except Exception as exc:
        raise RuntimeError("Pillow and pytesseract are required for OCR") from exc

    _configure_tesseract(pytesseract)

    img = Image.open(path)
    data = pytesseract.image_to_data(img, output_type=Output.DICT)

    parts: list[str] = []
    tokens: list[OcrToken] = []
    cursor = 0

    for idx, raw in enumerate(data.get("text", [])):
        word = (raw or "").strip()
        if not word:
            continue

        if parts:
            parts.append(" ")
            cursor += 1

        start = cursor
        parts.append(word)
        cursor += len(word)
        end = cursor

        tokens.append(
            OcrToken(
                text=word,
                start=start,
                end=end,
                left=int(data["left"][idx]),
                top=int(data["top"][idx]),
                width=int(data["width"][idx]),
                height=int(data["height"][idx]),
            )
        )

    return "".join(parts), tokens


def _extract_image_text(path: Path) -> str:
    text, _ = extract_image_text_with_tokens(path)
    return text


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_TYPES:
        raise ValueError(f"Unsupported file format: {suffix}")

    if suffix in {".txt", ".sql", ".csv"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if suffix == ".json":
        parsed = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
        return json.dumps(parsed, indent=2)
    if suffix == ".docx":
        return _extract_docx_text(path)
    if suffix == ".pdf":
        return _extract_pdf_text(path)
    if suffix == ".xlsx":
        return _extract_xlsx_text(path)
    if suffix in IMAGE_TYPES:
        return _extract_image_text(path)
    raise ValueError(f"Unsupported file format: {suffix}")


def extract_context_hints(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    hints: list[str] = []

    try:
        if suffix == ".json":
            data = json.loads(path.read_text(encoding="utf-8", errors="ignore"))

            def walk(obj):
                if isinstance(obj, dict):
                    for k, v in obj.items():
                        hints.append(str(k).lower())
                        walk(v)
                elif isinstance(obj, list):
                    for item in obj:
                        walk(item)

            walk(data)
        elif suffix == ".csv":
            first_line = path.read_text(encoding="utf-8", errors="ignore").splitlines()[:1]
            if first_line:
                hints.extend([c.strip().lower() for c in first_line[0].split(",")])
        elif suffix == ".sql":
            text = path.read_text(encoding="utf-8", errors="ignore").lower()
            hints.extend(re.findall(r"\b[a-z_][a-z0-9_]{2,}\b", text)[:300])
        elif suffix == ".xlsx":
            # sheet names and first-row strings are useful anchors for contextual linking
            try:
                from openpyxl import load_workbook
            except Exception:
                return []
            wb = load_workbook(filename=str(path), data_only=True, read_only=True)
            for ws in wb.worksheets[:10]:
                hints.append(ws.title.lower())
                try:
                    first = next(ws.iter_rows(min_row=1, max_row=1))
                except StopIteration:
                    continue
                for cell in first[:30]:
                    if isinstance(cell.value, str) and cell.value.strip():
                        hints.append(cell.value.strip().lower())
    except Exception:
        pass

    return list(dict.fromkeys(hints))


def _write_pdf(path: Path, sanitized_text: str) -> None:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except Exception as exc:
        raise RuntimeError("reportlab is required for PDF output") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path), pagesize=A4)
    _, height = A4
    y = height - 40
    c.setFont("Helvetica", 10)

    for line in sanitized_text.splitlines() or [""]:
        c.drawString(40, y, line[:150])
        y -= 14
        if y < 40:
            c.showPage()
            c.setFont("Helvetica", 10)
            y = height - 40

    c.save()


def _write_docx(path: Path, sanitized_text: str) -> None:
    content_types = """<?xml version=\"1.0\" encoding=\"UTF-8\"?>
<Types xmlns=\"http://schemas.openxmlformats.org/package/2006/content-types\">
  <Default Extension=\"rels\" ContentType=\"application/vnd.openxmlformats-package.relationships+xml\"/>
  <Default Extension=\"xml\" ContentType=\"application/xml\"/>
  <Override PartName=\"/word/document.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml\"/>
</Types>"""

    rels = """<?xml version=\"1.0\" encoding=\"UTF-8\"?>
<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">
  <Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument\" Target=\"word/document.xml\"/>
</Relationships>"""

    paragraphs = []
    for line in sanitized_text.splitlines() or [""]:
        paragraphs.append(f"<w:p><w:r><w:t xml:space=\"preserve\">{escape(line)}</w:t></w:r></w:p>")

    document_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<w:document xmlns:w=\"http://schemas.openxmlformats.org/wordprocessingml/2006/main\">"
        f"<w:body>{''.join(paragraphs)}</w:body></w:document>"
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("word/document.xml", document_xml)


def _write_xlsx(template_path: Path, output_path: Path, sanitized_text: str) -> None:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required for XLSX output") from exc

    lines, coords = _xlsx_to_lines(template_path)
    sanitized_lines = sanitized_text.splitlines()
    wb = load_workbook(filename=str(template_path))

    # Best-effort mapping: if the line count diverges, don't crash.
    if len(sanitized_lines) != len(lines):
        ws = wb.worksheets[0]
        ws["A1"].value = sanitized_text[:32000]
    else:
        for (sheet_name, cell_coord), new_val in zip(coords, sanitized_lines, strict=False):
            try:
                wb[sheet_name][cell_coord].value = new_val
            except Exception:
                continue

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def write_sanitized_output(path: Path, sanitized_text: str) -> None:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        _write_pdf(path, sanitized_text)
        return
    if suffix == ".docx":
        _write_docx(path, sanitized_text)
        return
    if suffix == ".xlsx":
        raise RuntimeError("Use write_sanitized_xlsx() for XLSX output")

    path.write_text(sanitized_text, encoding="utf-8")


def write_sanitized_xlsx(template_path: Path, output_path: Path, sanitized_text: str) -> None:
    _write_xlsx(template_path=template_path, output_path=output_path, sanitized_text=sanitized_text)


def as_download_stream(content: bytes):
    return io.BytesIO(content)
